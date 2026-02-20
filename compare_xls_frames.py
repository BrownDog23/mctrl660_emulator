import re
from typing import Optional, Dict

import openpyxl

XLSX_PATH = "LED-Side-Panels-Values.xlsx"
SHEET_NAME = "Tabelle1"


def norm_hex(s: Optional[str]) -> Optional[str]:
    """Normalizza una stringa esadecimale (rimuove spazi e caratteri non-hex)."""
    if s is None:
        return None
    s = str(s).strip().lower()
    if not s:
        return None
    s = re.sub(r"[^0-9a-f]", "", s)
    return s if s else None


def build_left_frame(value: int) -> bytes:
    """Frame FE (21 bytes) - LEFT, coerente con lo sniff."""
    if not (0 <= value <= 255):
        raise ValueError("value must be in 0..255")

    addr_low = (0x85 + 2 * value) & 0xFF

    fixed = bytes.fromhex(
        "55aa"      # header
        "0000"      # addr placeholder
        "fe"        # type
        "000104"    # param LEFT
        "ffff"
        "0100"
        "0100"
        "0002"
        "0100"
        "00"        # value placeholder (byte 18)
        "0000"      # crc placeholder (bytes 19-20)
    )

    b = bytearray(fixed)
    b[2] = 0x00
    b[3] = addr_low
    b[18] = value & 0xFF

    # CRC16 "lineare" osservato: base 0x58E0, +3*value, little-endian
    crc16 = (0x58E0 + 3 * value) & 0xFFFF
    b[19] = crc16 & 0xFF         # low
    b[20] = (crc16 >> 8) & 0xFF  # high

    return bytes(b)


def build_right_frame(value: int) -> bytes:
    """Frame FE (21 bytes) - RIGHT, coerente con lo sniff."""
    if not (0 <= value <= 255):
        raise ValueError("value must be in 0..255")

    addr_low = (0x86 + 2 * value) & 0xFF

    fixed = bytes.fromhex(
        "55aa"
        "0000"
        "fe"
        "000105"    # param RIGHT
        "ffff"
        "0100"
        "0100"
        "0002"
        "0100"
        "00"        # value placeholder
        "0000"      # crc placeholder
    )

    b = bytearray(fixed)
    b[2] = 0x00
    b[3] = addr_low
    b[18] = value & 0xFF

    # CRC16 osservato: base 0x58E2, +3*value, little-endian
    crc16 = (0x58E2 + 3 * value) & 0xFFFF
    b[19] = crc16 & 0xFF
    b[20] = (crc16 >> 8) & 0xFF

    return bytes(b)


def mismatch_kind(xls_hex: str, gen_hex: str) -> str:
    """
    Classifica mismatch:
    - LEN_DIFF: lunghezza diversa
    - ONLY_LAST_CRC_BYTE_DIFF: tutto uguale tranne ultimo byte (CRC high)
    - ONLY_VALUE_BYTE_DIFF: tutto uguale tranne byte value (pos 18) e CRC coerente col value del foglio? (semplificato)
    - REAL_DIFF: altro
    """
    if len(xls_hex) != len(gen_hex):
        return "LEN_DIFF"

    # differisce solo l'ultimo byte (2 hex chars)
    if xls_hex[:-2] == gen_hex[:-2] and xls_hex[-2:] != gen_hex[-2:]:
        return "ONLY_LAST_CRC_BYTE_DIFF"

    # differisce solo il byte value (pos 18 -> byte index 18 => hex indices 36..37)
    # NB: ogni byte = 2 hex chars. value byte start = 18*2 = 36, end = 38
    v_start, v_end = 36, 38
    if (
        xls_hex[:v_start] == gen_hex[:v_start]
        and xls_hex[v_end:] == gen_hex[v_end:]
        and xls_hex[v_start:v_end] != gen_hex[v_start:v_end]
    ):
        return "ONLY_VALUE_BYTE_DIFF"

    return "REAL_DIFF"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Foglio '{SHEET_NAME}' non trovato. Trovati: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h is not None}

    def col(name: str) -> int:
        key = name.strip().lower()
        if key not in header_map:
            raise RuntimeError(f"Colonna '{name}' non trovata. Intestazioni: {headers}")
        return header_map[key] + 1  # 1-based

    c_value = col("Value")
    c_left = col("Left")
    c_right = col("Right")

    stats: Dict[str, int] = {
        "TOTAL": 0,
        "MATCH": 0,
        "MISMATCH": 0,
        "LEN_DIFF": 0,
        "ONLY_LAST_CRC_BYTE_DIFF": 0,
        "ONLY_VALUE_BYTE_DIFF": 0,
        "REAL_DIFF": 0,
    }

    print("=== START COMPARE ===")

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=c_value).value
        if v is None:
            continue

        try:
            value_int = int(v)
        except Exception:
            continue

        left = norm_hex(ws.cell(row=r, column=c_left).value)
        right = norm_hex(ws.cell(row=r, column=c_right).value)

        # LEFT
        if left:
            stats["TOTAL"] += 1
            gen_left = build_left_frame(value_int).hex()

            if gen_left == left:
                stats["MATCH"] += 1
            else:
                stats["MISMATCH"] += 1
                kind = mismatch_kind(left, gen_left)
                stats[kind] += 1

                # stampa "compatta" per i casi ripetitivi
                if kind == "ONLY_LAST_CRC_BYTE_DIFF":
                    print(f"[ROW {r}] LEFT {kind} value={value_int} xls_last={left[-2:]} gen_last={gen_left[-2:]}")
                else:
                    print(f"[ROW {r}] LEFT {kind} value={value_int}")
                    print(f"  xls: {left}")
                    print(f"  gen: {gen_left}")

        # RIGHT
        if right:
            stats["TOTAL"] += 1
            gen_right = build_right_frame(value_int).hex()

            if gen_right == right:
                stats["MATCH"] += 1
            else:
                stats["MISMATCH"] += 1
                kind = mismatch_kind(right, gen_right)
                stats[kind] += 1

                if kind == "ONLY_LAST_CRC_BYTE_DIFF":
                    print(f"[ROW {r}] RIGHT {kind} value={value_int} xls_last={right[-2:]} gen_last={gen_right[-2:]}")
                else:
                    print(f"[ROW {r}] RIGHT {kind} value={value_int}")
                    print(f"  xls: {right}")
                    print(f"  gen: {gen_right}")

    print("\n=== SUMMARY ===")
    print("Frames compared:", stats["TOTAL"])
    print("Matches:", stats["MATCH"])
    print("Mismatches:", stats["MISMATCH"])
    print(" - LEN_DIFF:", stats["LEN_DIFF"])
    print(" - ONLY_LAST_CRC_BYTE_DIFF:", stats["ONLY_LAST_CRC_BYTE_DIFF"])
    print(" - ONLY_VALUE_BYTE_DIFF:", stats["ONLY_VALUE_BYTE_DIFF"])
    print(" - REAL_DIFF:", stats["REAL_DIFF"])

    if stats["REAL_DIFF"] == 0 and stats["LEN_DIFF"] == 0:
        print("\n✅ Formula e layout frame sono coerenti; i mismatch residui sono quasi certamente errori del foglio (CRC high o value).")
    else:
        print("\n⚠️ Ci sono mismatch REAL_DIFF/LEN_DIFF: potrebbe esserci più di una famiglia di frame o un campo interpretato male.")


if __name__ == "__main__":
    main()
